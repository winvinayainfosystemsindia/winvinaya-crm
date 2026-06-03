from typing import List, Optional, Dict, Any
from uuid import UUID
from datetime import datetime, timezone
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import HTTPException, status

from app.models.placement_mapping import PlacementMapping
from app.models.job_role import JobRole
from app.models.candidate import Candidate
from app.repositories.placement_mapping_repository import PlacementMappingRepository
from app.repositories.job_role_repository import JobRoleRepository
from app.repositories.candidate_repository import CandidateRepository
from app.schemas.placement_mapping import (
    PlacementMappingCreate, 
    CandidateMatchResult, 
    MatchMatchInfo,
    PlacementMappingBulkCreate,
    AIScoreRequest,
    AIScoreResponse,
    AIScoreResultItem,
)


class PlacementMappingService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repository = PlacementMappingRepository(db)
        self.job_role_repo = JobRoleRepository(db)
        self.candidate_repo = CandidateRepository(db)

    async def get_mapped_candidates(self, job_role_public_id: UUID) -> List[PlacementMapping]:
        job_role = await self.job_role_repo.get_by_public_id(job_role_public_id)
        if not job_role:
            raise HTTPException(status_code=404, detail="Job role not found")
        return await self.repository.get_by_job_role_active(job_role.id)

    async def map_candidate(self, mapping_in: PlacementMappingCreate, user_id: int) -> PlacementMapping:
        # Check if already mapped
        existing = await self.repository.get_by_candidate_and_job_role(
            mapping_in.candidate_id, mapping_in.job_role_id
        )
        if existing:
            raise HTTPException(
                status_code=400, detail="Candidate is already mapped to this job role"
            )

        mapping_data = {
            "candidate_id": mapping_in.candidate_id,
            "job_role_id": mapping_in.job_role_id,
            "match_score": mapping_in.match_score,
            "notes": mapping_in.notes,
            "mapped_by_id": user_id,
            "mapped_at": datetime.now(timezone.utc).replace(tzinfo=None),
            "ai_explanation": getattr(mapping_in, "ai_explanation", None),
            "score_source": getattr(mapping_in, "score_source", "rule_based") or "rule_based",
        }
        return await self.repository.create(mapping_data)

    async def bulk_map_candidates(
        self, bulk_mapping: PlacementMappingBulkCreate, user_id: int
    ) -> List[PlacementMapping]:
        """Map multiple candidates to a job role in one pass"""
        results = []
        for item in bulk_mapping.mappings:
            # Check if already mapped
            existing = await self.repository.get_by_candidate_and_job_role(
                item.candidate_id, bulk_mapping.job_role_id
            )
            if existing:
                continue # Skip if already mapped
            
            mapping_data = {
                "candidate_id": item.candidate_id,
                "job_role_id": bulk_mapping.job_role_id,
                "match_score": item.match_score,
                "notes": bulk_mapping.notes,
                "mapped_by_id": user_id,
                "mapped_at": datetime.now(timezone.utc).replace(tzinfo=None),
                "ai_explanation": getattr(item, "ai_explanation", None),
                "score_source": getattr(item, "score_source", "rule_based") or "rule_based",
            }
            mapping = await self.repository.create(mapping_data)
            results.append(mapping)
        return results

    async def unmap_candidate(self, candidate_id: int, job_role_id: int) -> bool:
        mapping = await self.repository.get_by_candidate_and_job_role(candidate_id, job_role_id)
        if not mapping:
            raise HTTPException(status_code=404, detail="Mapping not found")
        # Use hard delete since PlacementMapping doesn't support soft delete
        return await self.repository.delete(mapping.id, soft=False)

    async def get_matches_for_job_role(self, job_role_public_id: UUID, mapped_only: bool = False) -> List[CandidateMatchResult]:
        job_role = await self.job_role_repo.get_by_public_id(job_role_public_id)
        if not job_role:
            raise HTTPException(status_code=404, detail="Job role not found")

        # requirements: {"skills": [], "qualifications": [], "disability_preferred": []}
        reqs = job_role.requirements or {}
        
        def normalize_text(obj):
            if not obj: return []
            if isinstance(obj, str): return [obj.strip().lower()]
            if isinstance(obj, list): return [str(item).strip().lower() for item in obj if item]
            return []

        def normalize_qual(q_list):
            normalized = set()
            # Common aliases to bridge variations
            aliases = {
                "b.com": ["bcom", "b. com", "b com", "bachelor of commerce"],
                "b.e": ["be", "b. e", "b e", "bachelor of engineering"],
                "b.tech": ["btech", "b. tech", "b tech", "bachelor of technology"],
                "m.com": ["mcom", "m. com", "m com", "master of commerce"],
                "any graduation": ["any degree", "graduation", "any", "all", "graduation required", "any graduate"]
            }
            
            for q in q_list:
                q = q.lower().strip()
                if not q: continue
                normalized.add(q)
                # Add aliases
                for canonical, variations in aliases.items():
                    if q == canonical or q in variations:
                        normalized.add(canonical)
                        for v in variations: normalized.add(v)
            return normalized

        job_skills = set(normalize_text(reqs.get("skills")))
        job_quals = normalize_qual(normalize_text(reqs.get("qualifications")))
        job_disability = set(normalize_text(reqs.get("disability_preferred")))
        
        # Check for global match keywords early
        global_match_keywords = {"any graduation", "any degree", "graduation", "any", "all", "graduation required", "any graduate"}
        has_global_requirement = any(kw in job_quals for kw in global_match_keywords)

        # Fetch only placement-ready candidates: Screened (Completed) AND Counseled (Selected)
        candidates, _ = await self.candidate_repo.get_screened(
            limit=None, # Retrieve all records selected in counseling
            screening_status='Completed',
            counseling_status='selected'
        )
        
        # If mapped_only is true, we ONLY care about candidates who are ALREADY mapped to THIS job role.
        if mapped_only:
            # existing_mappings is fetched early, so we can filter immediately
            existing_mappings = await self.repository.get_by_job_role_active(job_role.id)
            mapped_candidate_ids = {m.candidate_id for m in existing_mappings}
            candidates = [c for c in candidates if c.id in mapped_candidate_ids]
        
        # We intentionally DO NOT eagerly load attendance, mock_interviews, or candidate_analyses here.
        # Those are only needed for AI scoring, which is handled in a separate endpoint.
        # Loading them here causes massive performance issues for the pipeline tabs.
        if candidates:
            from sqlalchemy import select
            from sqlalchemy.orm import selectinload
            from app.models.candidate import Candidate as CandidateModel
            cand_ids = [c.id for c in candidates]
            enriched_stmt = (
                select(CandidateModel)
                .where(CandidateModel.id.in_(cand_ids))
                .options(
                    selectinload(CandidateModel.screening),
                    selectinload(CandidateModel.counseling),
                )
            )
            enriched_result = await self.db.execute(enriched_stmt)
            enriched_map = {c.id: c for c in enriched_result.scalars().unique().all()}
            # Replace candidates list with enriched versions (preserving order)
            candidates = [enriched_map.get(c.id, c) for c in candidates]
        
        if not candidates:
            return []

        # Map of candidate_id -> (mapping_id, status)
        # existing_mappings was fetched earlier (either for mapped_only or we fetch it here if we didn't yet)
        if not mapped_only:
            existing_mappings = await self.repository.get_by_job_role_active(job_role.id)
            
        mapping_info = {m.candidate_id: (m.id, m.status) for m in existing_mappings}

        # Get active mappings for all fetched candidates to avoid N+1 query loop
        candidate_ids = [c.id for c in candidates]
        
        # Query active placement mappings in bulk using a lightweight core select
        from sqlalchemy import select, and_
        from app.models.company import Company
        
        stmt = (
            select(
                PlacementMapping.candidate_id,
                PlacementMapping.status,
                JobRole.id.label("job_role_id"),
                JobRole.title.label("job_role_title"),
                JobRole.status.label("job_role_status"),
                Company.name.label("company_name")
            )
            .join(JobRole, PlacementMapping.job_role_id == JobRole.id)
            .outerjoin(Company, JobRole.company_id == Company.id)
            .where(
                and_(
                    PlacementMapping.candidate_id.in_(candidate_ids),
                    PlacementMapping.is_active == True
                )
            )
        )
        bulk_result = await self.repository.db.execute(stmt)
        mapping_rows = bulk_result.all()
        
        # Group other mappings by candidate_id
        from collections import defaultdict
        mappings_by_candidate = defaultdict(list)
        for row in mapping_rows:
            mappings_by_candidate[row.candidate_id].append(row)

        results = []
        for candidate in candidates:
            # Skip candidate if they have an active placement mapping with placed statuses in a different job role
            # EXCEPT if they are already mapped to this current job role. If they are already mapped, we still
            # want them to show up in the Kanban pipeline so the recruiter knows their status.
            other_mappings = mappings_by_candidate.get(candidate.id, [])
            is_placed_elsewhere = False
            placed_elsewhere_info = None
            placed_statuses = {"joined", "offered", "offer_made", "offer_accepted"}
            for row in other_mappings:
                status_str = getattr(row.status, "value", str(row.status)).lower().strip()
                if row.job_role_id != job_role.id and status_str in placed_statuses:
                    is_placed_elsewhere = True
                    company_name = row.company_name or "Another Company"
                    placed_elsewhere_info = f"Placed at {company_name} as {row.job_role_title}"
                    break
            
            is_already_mapped = candidate.id in mapping_info

            if is_placed_elsewhere and not is_already_mapped:
                # If they are placed elsewhere and NOT mapped to this role, completely hide them
                # from the "Map Candidates" dialog.
                continue

            # 1. Skill Match (60%)
            candidate_skills = []
            
            def get_names_from_skills(obj):
                names = []
                if not obj: return names
                if isinstance(obj, dict):
                    # Handle categorized dict: {"technical": [], "soft": []}
                    names.extend(obj.get("technical") or [])
                    names.extend(obj.get("soft") or [])
                    # Handle flat dict if names are keys or something else? 
                    # Usually it's categorized per our models.
                elif isinstance(obj, list):
                    # Handle list of objects: [{"name": "Skill", "level": "..."}]
                    for item in obj:
                        if isinstance(item, dict):
                            name = item.get("name") or item.get("skill")
                            if name: names.append(name)
                        elif isinstance(item, str):
                            names.append(item)
                return names

            # Extract from screening
            if candidate.screening and candidate.screening.skills:
                candidate_skills.extend(get_names_from_skills(candidate.screening.skills))
            
            # Extract from counseling
            if candidate.counseling and candidate.counseling.skills:
                candidate_skills.extend(get_names_from_skills(candidate.counseling.skills))
            
            c_skills_set = set(s.lower() for s in candidate_skills)
            skill_hits = job_skills.intersection(c_skills_set)
            skill_score = 0.0
            skill_detail = "No skills specified in job role" if not job_skills else "No matching skills"
            
            if job_skills:
                skill_ratio = len(skill_hits) / len(job_skills)
                skill_score = skill_ratio * 60.0
                skill_detail = f"Matched {len(skill_hits)}/{len(job_skills)} skills: {', '.join(skill_hits)}"

            # 2. Qualification Match (20%)
            raw_c_quals = []
            if candidate.education_details and 'degrees' in candidate.education_details:
                for deg in candidate.education_details['degrees']:
                    if isinstance(deg, dict):
                        # Extract both degree name and major to catch variations like B.Com (Computer Science)
                        c_deg = (deg.get('degree_name') or deg.get('degree') or deg.get('name') or '').lower().strip()
                        c_major = (deg.get('major') or deg.get('specialization') or '').lower().strip()
                        if c_deg: raw_c_quals.append(c_deg)
                        if c_major: raw_c_quals.append(c_major)
            
            # Normalize candidate qualifications using the same alias logic
            candidate_quals = normalize_qual(raw_c_quals)
            
            # Match logic: True if global match OR any candidate qualification matches job requirements
            qual_match = has_global_requirement or any(q in job_quals for q in candidate_quals if q)
            
            qual_score = 20.0 if qual_match else 0.0
            qual_detail = "Matched qualification" if qual_match else "Qualification does not match"
            
            if not job_quals or has_global_requirement:
                qual_score = 20.0
                qual_detail = "Any graduation/qualification acceptable"

            # 3. Disability Match (20%)
            candidate_disability = ""
            if candidate.disability_details:
                candidate_disability = (candidate.disability_details.get("disability_type") or "").lower()
            
            dis_match = candidate_disability in job_disability
            dis_score = 20.0 if dis_match else 0.0
            dis_detail = f"Matched disability type: {candidate_disability}" if dis_match else "Disability type mismatch"
            if not job_disability:
                dis_score = 20.0
                dis_detail = "No disability preference specified"

            total_score = skill_score + qual_score + dis_score
            
            # Count and fetch other mappings for this candidate (Only Active roles) - fast dictionary lookup
            other_mappings = mappings_by_candidate.get(candidate.id, [])
            other_role_names = [
                row.job_role_title for row in other_mappings 
                if row.job_role_id != job_role.id and getattr(row.job_role_status, "value", str(row.job_role_status)).lower() == "active"
            ]
            other_count = len(other_role_names)

            # Extract top-level qualification and disability
            main_qual = ""
            if candidate.education_details and 'degrees' in candidate.education_details and candidate.education_details['degrees']:
                 deg = candidate.education_details['degrees'][0]
                 main_qual = deg.get('degree_name') or deg.get('degree', '')

            main_disability = ""
            if candidate.disability_details:
                main_disability = candidate.disability_details.get("disability_type")

            # Extract year of experience
            year_of_exp = None
            if candidate.work_experience and isinstance(candidate.work_experience, dict):
                year_of_exp = candidate.work_experience.get("year_of_experience")

            # Extract registration source (registration_type) from candidate's other field
            reg_source = None
            if candidate.other and isinstance(candidate.other, dict):
                reg_source = candidate.other.get("registration_type")
            if not reg_source:
                reg_source = "Registered"

            results.append(
                CandidateMatchResult(
                    public_id=candidate.public_id,
                    candidate_id=candidate.id,
                    name=candidate.name,
                    match_score=round(total_score, 2),
                    disability=main_disability,
                    qualification=main_qual,
                    skills=candidate_skills[:5],  # Top 5 skills
                    skill_match=MatchMatchInfo(is_match=len(skill_hits) > 0, details=skill_detail),
                    qualification_match=MatchMatchInfo(is_match=qual_match or not job_quals, details=qual_detail),
                    disability_match=MatchMatchInfo(is_match=dis_match or not job_disability, details=dis_detail),
                    other_mappings_count=other_count,
                    other_mappings=other_role_names,
                    is_already_mapped=is_already_mapped,
                    is_placed_elsewhere=is_placed_elsewhere,
                    placed_elsewhere_info=placed_elsewhere_info,
                    status=mapping_info[candidate.id][1] if candidate.id in mapping_info else None,
                    mapping_id=mapping_info[candidate.id][0] if candidate.id in mapping_info else None,
                    year_of_experience=year_of_exp,
                    source_of_info=reg_source
                )
            )

        # Sort by match score descending
        results.sort(key=lambda x: x.match_score, reverse=True)
        return results

    async def ai_score_candidates(
        self, job_role_public_id: UUID, request: AIScoreRequest
    ) -> AIScoreResponse:
        """
        AI-score a specific set of candidates against a job role.
        Called on-demand when the user clicks "Score with AI".

        Returns AIScoreResponse with a dict of str(candidate_id) -> AIScoreResultItem.
        """
        from sqlalchemy import select
        from sqlalchemy.orm import selectinload
        from app.models.candidate import Candidate as CandidateModel
        from app.ai.services.placement_scoring_service import PlacementScoringService

        job_role = await self.job_role_repo.get_by_public_id(job_role_public_id)
        if not job_role:
            raise HTTPException(status_code=404, detail="Job role not found")

        if not request.candidate_ids:
            return AIScoreResponse(scores={})

        # Fetch candidates with all relationships needed for scoring
        stmt = (
            select(CandidateModel)
            .where(CandidateModel.id.in_(request.candidate_ids))
            .options(
                selectinload(CandidateModel.screening),
                selectinload(CandidateModel.counseling),
                selectinload(CandidateModel.attendance),
                selectinload(CandidateModel.mock_interviews),
            )
        )
        result = await self.db.execute(stmt)
        candidates = list(result.scalars().unique().all())

        if not candidates:
            return AIScoreResponse(scores={})

        # Run AI scoring with bounded concurrency
        scoring_service = PlacementScoringService(self.db)
        scores_map = await scoring_service.score_candidates(job_role, candidates)

        # Build response — keys must be str for JSON serialization
        scores_out: Dict[str, AIScoreResultItem] = {}
        for cand_id, data in scores_map.items():
            scores_out[str(cand_id)] = AIScoreResultItem(
                score=data.get("score"),
                explanation=data.get("explanation"),
                recommendation=data.get("recommendation"),
                score_source=data.get("score_source", "rule_based"),
            )

        return AIScoreResponse(scores=scores_out)

    async def export_placement_mappings(
        self,
        job_role_public_id: UUID,
        current_user: Any,
        columns: Optional[str] = None
    ) -> bool:
        """Fetch all mapped candidates for a job role, generate Excel, and email to user"""
        import json
        import io
        from datetime import datetime, date
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from app.utils.email import send_export_email

        mappings = await self.get_mapped_candidates(job_role_public_id)
        
        job_role_title = mappings[0].job_role.title if mappings else "Placement"
        company_name = mappings[0].job_role.company.name if mappings and getattr(mappings[0].job_role, "company", None) else ""
        
        column_defs = []
        if columns:
            try:
                column_defs = json.loads(columns)
            except Exception:
                column_defs = []
        
        if not column_defs:
            column_defs = [
                {"id": "name", "label": "Candidate Name"},
                {"id": "email", "label": "Email"},
                {"id": "phone", "label": "Phone"},
                {"id": "status", "label": "Placement Status"}
            ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Placement Report"
        for col_num, col_def in enumerate(column_defs, 1):
            cell = ws.cell(row=1, column=col_num, value=col_def["label"])
            cell.font = Font(bold=True)
            
        for row_num, mapping in enumerate(mappings, 2):
            c = mapping.candidate
            for col_num, col_def in enumerate(column_defs, 1):
                col_id = col_def["id"]
                val = ""
                
                try:
                    if col_id == "mapped_company":
                        val = company_name
                    elif col_id == "status":
                        val = getattr(mapping.status, "value", str(mapping.status))
                    elif col_id == "batch_tag":
                        allocs = getattr(c, "allocations", [])
                        if allocs and len(allocs) > 0 and getattr(allocs[0], "batch", None):
                            val = getattr(allocs[0].batch, "batch_tag", "")
                        else:
                            val = ""
                    elif col_id == "registration_type":
                        val = (c.other or {}).get("registration_type", "Registered")
                    elif col_id == "disability_type":
                        val = (c.disability_details or {}).get("disability_type", "")
                    elif col_id == "disability_percentage":
                        val = (c.disability_details or {}).get("disability_percentage", "")
                    elif col_id == "education_level":
                        edu = c.education_details or {}
                        degrees = edu.get("degrees", [])
                        val = degrees[0].get("degree_name") if degrees else ""
                    elif col_id == "specialization":
                        edu = c.education_details or {}
                        degrees = edu.get("degrees", [])
                        val = degrees[0].get("specialization") if degrees else ""
                    elif col_id == "year_of_passing":
                        edu = c.education_details or {}
                        degrees = edu.get("degrees", [])
                        val = degrees[0].get("year_of_passing") if degrees else ""
                    elif col_id == "screening_status":
                        val = c.screening.status if getattr(c, "screening", None) else "Pending"
                    elif col_id == "consent_status":
                        val = c.screening.consent_status if getattr(c, "screening", None) else ""
                    elif col_id == "screening_date":
                        val = c.screening.created_at if getattr(c, "screening", None) else ""
                    elif col_id == "screened_by_name":
                        val = (c.screening.screened_by.full_name or c.screening.screened_by.username) if getattr(c, "screening", None) and getattr(c.screening, "screened_by", None) else ""
                    elif col_id == "counseling_status":
                        val = c.counseling.status if getattr(c, "counseling", None) else ""
                    elif col_id == "counseling_date":
                        val = c.counseling.counseling_date if getattr(c, "counseling", None) else ""
                    elif col_id == "counselor_name":
                        val = (c.counseling.counselor.full_name or c.counseling.counselor.username) if getattr(c, "counseling", None) and getattr(c.counseling, "counselor", None) else ""
                    elif col_id == "is_experienced":
                        val = (c.work_experience or {}).get("is_experienced", False)
                    elif col_id == "year_of_experience":
                        val = (c.work_experience or {}).get("year_of_experience", "")
                    elif col_id == "currently_employed":
                        val = (c.work_experience or {}).get("currently_employed", False)
                    elif col_id == "suitable_job_roles":
                        val = (c.counseling.others or {}).get("suitable_job_roles", []) if getattr(c, "counseling", None) else []
                    elif col_id == "source_of_info":
                        val = (c.screening.others or {}).get("source_of_info", "") if getattr(c, "screening", None) else ""
                    elif col_id == "family_annual_income":
                        val = (c.screening.others or {}).get("family_annual_income", "") if getattr(c, "screening", None) else ""
                    elif col_id == "screening_comments":
                        val = (c.screening.others or {}).get("reason", "") if getattr(c, "screening", None) else ""
                    elif col_id == "skills" and getattr(c, "counseling", None):
                        val = c.counseling.skills
                    elif col_id == "workexperience" and getattr(c, "counseling", None):
                        val = c.counseling.workexperience
                    elif col_id == "questions" and getattr(c, "counseling", None):
                        val = c.counseling.questions
                    elif col_id.startswith("screening_others."):
                        field_name = col_id.replace("screening_others.", "")
                        val = (c.screening.others or {}).get(field_name, "") if getattr(c, "screening", None) else ""
                    elif col_id.startswith("counseling_others."):
                        field_name = col_id.replace("counseling_others.", "")
                        val = (c.counseling.others or {}).get(field_name, "") if getattr(c, "counseling", None) else ""
                    elif "." in col_id:
                        parts = col_id.split(".")
                        obj = c
                        for part in parts:
                            if obj is None: break
                            if isinstance(obj, dict): obj = obj.get(part, "")
                            else: obj = getattr(obj, part, None)
                        val = obj
                    else:
                        val = getattr(c, col_id, "")
                        if (val is None or val == "") and getattr(c, "screening", None) and hasattr(c.screening, col_id):
                            val = getattr(c.screening, col_id, "")
                        if (val is None or val == "") and getattr(c, "counseling", None) and hasattr(c.counseling, col_id):
                            val = getattr(c.counseling, col_id, "")
                    
                    if isinstance(val, (datetime, date)):
                        val = val.strftime('%Y-%m-%d')
                    elif isinstance(val, bool):
                        val = "Yes" if val else "No"
                    elif val is None:
                        val = ""
                    elif isinstance(val, (dict, list)):
                        if col_id == "family_details" and isinstance(val, list):
                            val = "; ".join([f"{f.get('relation')}: {f.get('name')} ({f.get('occupation', 'N/A')})" for f in val])
                        elif col_id == "skills" and isinstance(val, list):
                            val = ", ".join([f"{s.get('name')} ({s.get('level')})" for s in val])
                        elif col_id == "workexperience" and isinstance(val, list):
                            val = ", ".join([f"{w.get('job_title')} at {w.get('company')}" for w in val])
                        elif col_id == "questions" and isinstance(val, list):
                            val = " | ".join([f"Q: {q.get('question')} A: {q.get('answer')}" for q in val])
                        else:
                            val = json.dumps(val)
                except Exception:
                    val = "Error"
                
                ws.cell(row=row_num, column=col_num, value=str(val) if val is not None else "")

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        report_name = f"Placement Report - {job_role_title}"
        filename = f"Placement_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return await send_export_email(
            to_email=current_user.email,
            user_name=current_user.full_name or current_user.username,
            report_name=report_name,
            file_content=excel_buffer.getvalue(),
            filename=filename
        )
