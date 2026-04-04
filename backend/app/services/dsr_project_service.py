"""DSR Project Service"""

from typing import List, Optional, Tuple
from uuid import UUID
import io
from fastapi import HTTPException, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.dsr_project import DSRProject, DSRProjectType
from app.models.user import User, UserRole
from app.schemas.dsr_project import (
    DSRProjectCreate, 
    DSRProjectUpdate, 
    DSRProjectImportResult,
    TrainingProjectSummary,
    TrainingProjectSubjectSummary
)
from app.repositories.dsr_project_repository import DSRProjectRepository
from app.repositories.user_repository import UserRepository
from app.repositories.dsr_entry_repository import DSREntryRepository
from app.repositories.dsr_activity_repository import DSRActivityRepository
from app.repositories.training_batch_repository import TrainingBatchRepository
from app.services.training_project_sync_service import TrainingProjectSyncService


def _require_manager_or_admin(current_user: User) -> None:
    if current_user.role not in (UserRole.ADMIN, UserRole.MANAGER):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Managers or Admins can manage projects",
        )


def _require_admin(current_user: User) -> None:
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can perform this action",
        )


class DSRProjectService:

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = DSRProjectRepository(db)
        self.user_repo = UserRepository(db)
        self.dsr_repo = DSREntryRepository(db)
        self.activity_repo = DSRActivityRepository(db)
        self.batch_repo = TrainingBatchRepository(db)
        self.sync_service = TrainingProjectSyncService(db)

    async def create_project(self, data: DSRProjectCreate, current_user: User) -> DSRProject:
        _require_manager_or_admin(current_user)
        # Resolve owner
        owner = await self.user_repo.get_by_fields(public_id=data.owner_user_public_id)
        if not owner:
            raise HTTPException(status_code=404, detail="Owner user not found")
        owner_user = owner[0] if owner else None
        if not owner_user:
            raise HTTPException(status_code=404, detail="Owner user not found")

        # Resolve multiple batches if training project
        linked_batches = []
        linked_batch_id = None # legacy support
        
        if data.project_type == "training":
            if data.linked_batch_public_ids:
                for bpid in data.linked_batch_public_ids:
                    batch = await self.batch_repo.get_by_public_id(str(bpid))
                    if batch:
                        linked_batches.append(batch)
            elif data.linked_batch_public_id:
                # Backwards compatibility
                batch = await self.batch_repo.get_by_public_id(str(data.linked_batch_public_id))
                if batch:
                    linked_batches.append(batch)
                    linked_batch_id = batch.id

        project_data = {
            "name": data.name,
            "owner_id": owner_user.id,
            "created_by": current_user.id,
            "is_active": data.is_active,
            "project_type": data.project_type,
            "linked_batch_id": linked_batch_id,
            "others": data.others,
        }
        project = await self.repo.create(project_data)
        
        # Link batches (many-to-many)
        if linked_batches:
            project.linked_batches = linked_batches
            await self.db.flush()
        
        # Trigger sync if training project
        if project.project_type == DSRProjectType.TRAINING and (project.linked_batches or project.linked_batch_id):
            await self.sync_service.sync_activities_for_project(project.id)
            
        return project

    async def update_project(
        self, public_id: UUID, data: DSRProjectUpdate, current_user: User
    ) -> DSRProject:
        _require_manager_or_admin(current_user)
        project = await self._get_or_404(public_id)

        update_data: dict = data.model_dump(exclude_unset=True)

        if "owner_user_public_id" in update_data:
            owner_pid = update_data.pop("owner_user_public_id")
            owner = await self.user_repo.get_by_fields(public_id=owner_pid)
            if not owner:
                raise HTTPException(status_code=404, detail="New owner user not found")
            update_data["owner_id"] = owner[0].id

        if "linked_batch_public_ids" in update_data:
            batch_pids = update_data.pop("linked_batch_public_ids")
            if batch_pids is not None:
                batches = []
                for bpid in batch_pids:
                    batch = await self.batch_repo.get_by_public_id(str(bpid))
                    if batch:
                        batches.append(batch)
                project.linked_batches = batches
            else:
                project.linked_batches = []
        elif "linked_batch_public_id" in update_data:
            batch_pid = update_data.pop("linked_batch_public_id")
            if batch_pid:
                batch = await self.batch_repo.get_by_public_id(str(batch_pid))
                if batch:
                    project.linked_batches = [batch]
                    update_data["linked_batch_id"] = batch.id
            else:
                project.linked_batches = []
                update_data["linked_batch_id"] = None

        updated = await self.repo.update(project.id, update_data)
        
        # Re-sync if it's a training project or if the batches were changed
        if updated.project_type == DSRProjectType.TRAINING and (updated.linked_batches or updated.linked_batch_id):
            await self.sync_service.sync_activities_for_project(updated.id)
            
        return await self._get_or_404(public_id)

    async def delete_project(self, public_id: UUID, current_user: User) -> bool:
        _require_manager_or_admin(current_user)
        project = await self._get_or_404(public_id)

        # Managers can only delete projects they created
        if current_user.role == UserRole.MANAGER and project.created_by != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Managers can only delete projects they created",
            )
        
        # Check if project is used in any DSR entries
        usage_count = await self.dsr_repo.count_references(project_public_id=project.public_id)
        if usage_count > 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Project '{project.name}' cannot be deleted because it is referenced in {usage_count} DSR entry/entries. Please deactivate it instead.",
            )
            
        return await self.repo.delete(project.id)

    async def get_projects(
        self,
        current_user: User,
        skip: int = 0,
        limit: int = 100,
        active_only: bool = False,
        assigned_to_public_id: Optional[UUID] = None,
        search: Optional[str] = None,
    ) -> Tuple[List[DSRProject], int]:
        assigned_to_id = None
        is_privileged = current_user.role in (UserRole.ADMIN, UserRole.MANAGER)
        
        if assigned_to_public_id:
            user = await self.user_repo.get_by_fields(public_id=assigned_to_public_id)
            if user:
                assigned_to_id = user[0].id
        
        owned_or_assigned_to = None
        if not is_privileged:
            owned_or_assigned_to = current_user.id

        return await self.repo.get_multi_paginated(
            skip=skip, 
            limit=limit, 
            active_only=active_only, 
            assigned_to=assigned_to_id, 
            owned_or_assigned_to=owned_or_assigned_to,
            search=search
        )

    async def get_project(self, public_id: UUID, current_user: User) -> DSRProject:
        project = await self._get_or_404(public_id)
        
        # Access control
        if current_user.role not in (UserRole.ADMIN, UserRole.MANAGER):
            if project.owner_id != current_user.id:
                # Check if assigned to any activities in this project
                items, total = await self.repo.get_multi_paginated(
                    limit=1,
                    owned_or_assigned_to=current_user.id,
                    search=project.name # A bit hacky but works since we know the project exists
                )
                # Verify if the returned project matches our public_id
                if not any(p.public_id == public_id for p in items):
                     raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="You do not have access to this project",
                    )
                    
        return project

    async def import_from_excel(
        self, file: UploadFile, current_user: User
    ) -> DSRProjectImportResult:
        """
        Bulk import projects from an Excel file.
        Expected columns: name, owner_email, is_active (optional), others (optional JSON)
        """
        _require_manager_or_admin(current_user)

        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl not installed. Add it to requirements.txt.",
            )

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        required = {"name", "owner_email"}
        missing_headers = required - set(headers)
        if missing_headers:
            raise HTTPException(
                status_code=422,
                detail=f"Excel missing required columns: {missing_headers}",
            )

        result = DSRProjectImportResult(total_rows=0, created=0, skipped=0, errors=[])

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            result.total_rows += 1

            name = str(row_data.get("name", "")).strip()
            owner_email = str(row_data.get("owner_email", "")).strip()

            if not name or not owner_email:
                result.skipped += 1
                result.errors.append({"row": row_idx, "error": "Missing name or owner_email"})
                continue

            # Check duplicate
            existing = await self.repo.get_by_name(name)
            if existing:
                result.skipped += 1
                result.errors.append({"row": row_idx, "error": f"Project '{name}' already exists"})
                continue

            owner = await self.user_repo.get_by_email(owner_email)
            if not owner:
                result.skipped += 1
                result.errors.append({"row": row_idx, "error": f"User with email '{owner_email}' not found"})
                continue

            is_active_val = row_data.get("is_active")
            is_active = True if is_active_val is None else bool(is_active_val)

            await self.repo.create({
                "name": name,
                "owner_id": owner.id,
                "created_by": current_user.id,
                "is_active": is_active,
                "others": None,
            })
            result.created += 1

        return result

    async def get_import_template(self) -> io.BytesIO:
        """Generate a blank Excel template for project import."""
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl not installed. Add it to requirements.txt.",
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Import Template"

        headers = ["name", "owner_email", "is_active"]
        ws.append(headers)

        # Add a sample row
        ws.append([
            "Sample Project", 
            "admin@example.com", 
            "TRUE"
        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def get_training_summary(self, public_id: UUID, current_user: User) -> TrainingProjectSummary:
        """Get a detailed planned vs actual summary for a training project"""
        project = await self._get_or_404(public_id)
        
        if project.project_type != DSRProjectType.TRAINING:
            raise HTTPException(status_code=400, detail="This project is not a training project")
            
        activities = await self.activity_repo.get_activities_for_project(project.id)
        
        total_planned = 0.0
        total_actual = 0.0
        subjects = []
        
        for act in activities:
            planned = act.estimated_hours or 0.0
            actual = act.total_actual_hours or 0.0
            total_planned += planned
            total_actual += actual
            
            trainer = act.assigned_users[0] if act.assigned_users else None
            
            subjects.append(TrainingProjectSubjectSummary(
                name=act.name,
                trainer_name=trainer.full_name or trainer.username if trainer else "Unassigned",
                trainer_public_id=trainer.public_id if trainer else None,
                planned_hours=round(planned, 2),
                actual_hours=round(actual, 2),
                completion_percentage=round((actual / planned * 100) if planned > 0 else 0, 1),
                status=act.status.value
            ))
            
        if project.linked_batches:
            batch_names = [b.batch_name for b in project.linked_batches]
            batch_display_name = ", ".join(batch_names)
        else:
            batch_display_name = project.linked_batch.batch_name if project.linked_batch else None

        return TrainingProjectSummary(
            project_id=project.id,
            project_public_id=project.public_id,
            project_name=project.name,
            batch_name=batch_display_name,
            total_planned_hours=round(total_planned, 2),
            total_actual_hours=round(total_actual, 2),
            overall_completion_percentage=round((total_actual / total_planned * 100) if total_planned > 0 else 0, 1),
            subjects=subjects
        )

    async def sync_training_project(self, public_id: UUID, current_user: User) -> bool:
        """Manually trigger a sync for a training project"""
        _require_manager_or_admin(current_user)
        project = await self._get_or_404(public_id)
        
        if project.project_type == DSRProjectType.TRAINING and (project.linked_batches or project.linked_batch_id):
            await self.sync_service.sync_activities_for_project(project.id)
            return True
        return False

    async def _get_or_404(self, public_id: UUID) -> DSRProject:
        project = await self.repo.get_by_public_id(public_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        return project
