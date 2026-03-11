"""DSR Activity Service"""

from typing import List, Optional, Tuple
from uuid import UUID
import io
from datetime import date
from fastapi import HTTPException, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.dsr_activity import DSRActivity, DSRActivityStatus
from app.models.user import User, UserRole
from app.schemas.dsr_activity import DSRActivityCreate, DSRActivityUpdate, DSRActivityImportResult
from app.repositories.dsr_activity_repository import DSRActivityRepository
from app.repositories.dsr_project_repository import DSRProjectRepository
from app.repositories.dsr_entry_repository import DSREntryRepository
from app.repositories.user_repository import UserRepository



def _require_admin(current_user: User) -> None:
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can perform this action",
        )


class DSRActivityService:

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = DSRActivityRepository(db)
        self.project_repo = DSRProjectRepository(db)
        self.dsr_repo = DSREntryRepository(db)

    async def _check_project_ownership(self, project_id: int, current_user: User) -> None:
        """Raise 403 if current_user is not the project owner and not an admin."""
        if current_user.role in (UserRole.ADMIN, UserRole.MANAGER):
            return
        project = await self.project_repo.get(project_id)
        if not project or project.owner_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only the project owner or Admin can manage activities for this project",
            )

    async def create_activity(self, data: DSRActivityCreate, current_user: User) -> DSRActivity:
        project = await self.project_repo.get_by_public_id(data.project_public_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        await self._check_project_ownership(project.id, current_user)

        assigned_users = []
        if data.assigned_user_public_ids:
            user_repo = UserRepository(self.db)
            for upid in data.assigned_user_public_ids:
                user_list = await user_repo.get_by_fields(public_id=upid)
                if not user_list:
                    raise HTTPException(status_code=404, detail=f"User with ID {upid} not found")
                assigned_users.append(user_list[0])

        activity_data = {
            "project_id": project.id,
            "name": data.name,
            "description": data.description,
            "start_date": data.start_date,
            "end_date": data.end_date,
            "status": data.status,
            "assigned_users": assigned_users,
            "is_active": data.is_active,
            "others": data.others,
        }
        return await self.repo.create(activity_data)

    async def update_activity(
        self, public_id: UUID, data: DSRActivityUpdate, current_user: User
    ) -> DSRActivity:
        activity = await self._get_or_404(public_id)
        await self._check_project_ownership(activity.project_id, current_user)

        update_data = data.model_dump(exclude_unset=True)

        if "assigned_user_public_ids" in update_data:
            assigned_ids = update_data.pop("assigned_user_public_ids")
            if assigned_ids is not None:
                user_repo = UserRepository(self.db)
                assigned_users = []
                for upid in assigned_ids:
                    user_list = await user_repo.get_by_fields(public_id=upid)
                    if not user_list:
                        raise HTTPException(status_code=404, detail=f"User with ID {upid} not found")
                    assigned_users.append(user_list[0])
                activity.assigned_users = assigned_users
            else:
                activity.assigned_users = []

        # Validate date consistency after merge
        merged_start = update_data.get("start_date", activity.start_date)
        merged_end = update_data.get("end_date", activity.end_date)
        if merged_end < merged_start:
            raise HTTPException(status_code=422, detail="end_date must be on or after start_date")

        # Set other fields
        for key, value in update_data.items():
            setattr(activity, key, value)

        await self.db.flush()
        return await self._get_or_404(public_id)

    async def delete_activity(self, public_id: UUID, current_user: User) -> bool:
        activity = await self._get_or_404(public_id)
        await self._check_project_ownership(activity.project_id, current_user)
        
        # Check if activity is used in any DSR entries
        usage_count = await self.dsr_repo.count_references(activity_public_id=activity.public_id)
        if usage_count > 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Activity '{activity.name}' cannot be deleted because it is referenced in {usage_count} DSR entry/entries. Please deactivate it instead.",
            )
            
        return await self.repo.delete(activity.id)

    async def get_activity(self, public_id: UUID) -> DSRActivity:
        return await self._get_or_404(public_id)

    async def get_activities(
        self,
        current_user: User,
        skip: int = 0,
        limit: int = 100,
        project_public_id: Optional[UUID] = None,
        status: Optional[DSRActivityStatus] = None,
        active_only: bool = False,
        assigned_to_public_id: Optional[UUID] = None,
        search: Optional[str] = None,
    ) -> Tuple[List[DSRActivity], int]:
        project_id = None
        if project_public_id:
            project = await self.project_repo.get_by_public_id(project_public_id)
            if not project:
                raise HTTPException(status_code=404, detail="Project not found")
            project_id = project.id

        assigned_to_id = None
        is_privileged = current_user.role in (UserRole.ADMIN, UserRole.MANAGER)

        if assigned_to_public_id and not is_privileged:
            user_repo = UserRepository(self.db)
            user = await user_repo.get_by_fields(public_id=assigned_to_public_id)
            if user:
                assigned_to_id = user[0].id

        return await self.repo.get_multi_paginated(
            skip=skip,
            limit=limit,
            project_id=project_id,
            status=status,
            active_only=active_only,
            assigned_to=assigned_to_id,
            search=search,
        )

    async def import_from_excel(
        self, file: UploadFile, current_user: User, project_public_id: Optional[UUID] = None
    ) -> DSRActivityImportResult:
        """
        Bulk import activities from Excel.
        Expected columns: project_name, name, description, start_date (YYYY-MM-DD),
                          end_date (YYYY-MM-DD), status (optional)
        If project_public_id is provided, it overrides project_name column and 
        makes it optional.
        Only imports activities for projects that current_user owns (or all if Admin).
        """
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed.")

        # Pre-fetch project if ID provided
        forced_project = None
        if project_public_id:
            forced_project = await self.project_repo.get_by_public_id(project_public_id)
            if not forced_project:
                raise HTTPException(status_code=404, detail="Target project not found")
            # Ownership check for forced project
            if current_user.role not in (UserRole.ADMIN, UserRole.MANAGER) and forced_project.owner_id != current_user.id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="You do not own the target project",
                )

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        required = {"name", "start_date", "end_date"}
        if not forced_project:
            required.add("project_name")
            
        if missing := required - set(headers):
            raise HTTPException(status_code=422, detail=f"Excel missing required columns: {missing}")

        result = DSRActivityImportResult(total_rows=0, created=0, skipped=0, errors=[])

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            result.total_rows += 1

            project_name = str(row_data.get("project_name", "")).strip()
            name = str(row_data.get("name", "")).strip()
            raw_start = row_data.get("start_date")
            raw_end = row_data.get("end_date")

            if not all([name, raw_start, raw_end]) or (not forced_project and not project_name):
                result.skipped += 1
                result.errors.append({"row": row_idx, "error": "Missing required fields"})
                continue

            if forced_project:
                project = forced_project
            else:
                project = await self.project_repo.get_by_name(project_name)
                if not project:
                    result.skipped += 1
                    result.errors.append({"row": row_idx, "error": f"Project '{project_name}' not found"})
                    continue

                # Ownership check
                if current_user.role not in (UserRole.ADMIN, UserRole.MANAGER) and project.owner_id != current_user.id:
                    result.skipped += 1
                    result.errors.append({
                        "row": row_idx,
                        "error": f"You are not the owner of project '{project_name}'",
                    })
                    continue

            try:
                start_date = raw_start if isinstance(raw_start, date) else date.fromisoformat(str(raw_start))
                end_date = raw_end if isinstance(raw_end, date) else date.fromisoformat(str(raw_end))
            except ValueError:
                result.skipped += 1
                result.errors.append({"row": row_idx, "error": "Invalid date format (use YYYY-MM-DD)"})
                continue

            if end_date < start_date:
                result.skipped += 1
                result.errors.append({"row": row_idx, "error": "end_date must be >= start_date"})
                continue

            raw_status = str(row_data.get("status", "planned")).strip().lower()
            try:
                act_status = DSRActivityStatus(raw_status)
            except ValueError:
                act_status = DSRActivityStatus.PLANNED

            await self.repo.create({
                "project_id": project.id,
                "name": name,
                "description": str(row_data.get("description", "")).strip() or None,
                "start_date": start_date,
                "end_date": end_date,
                "status": act_status,
                "is_active": True,
                "others": None,
            })
            result.created += 1

        return result

    async def get_import_template(self) -> io.BytesIO:
        """Generate a blank Excel template for activity import."""
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activity Import Template"

        headers = ["project_name", "name", "description", "start_date", "end_date", "status"]
        ws.append(headers)

        # Add a sample row
        ws.append([
            "Sample Project", 
            "Design UI", 
            "Create mockups for main dashboard", 
            "2024-01-01", 
            "2024-01-15", 
            "planned"
        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def export_activities(self, project_public_id: UUID) -> io.BytesIO:
        """Export all activities for a project to Excel."""
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed.")

        project = await self.project_repo.get_by_public_id(project_public_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        activities, _ = await self.repo.get_multi_paginated(
            skip=0, limit=1000, project_id=project.id
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Activities - {project.name}"

        headers = ["project_name", "name", "description", "start_date", "end_date", "status"]
        ws.append(headers)

        for act in activities:
            ws.append([
                project.name,
                act.name,
                act.description or "",
                act.start_date.isoformat(),
                act.end_date.isoformat(),
                act.status.value
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def _get_or_404(self, public_id: UUID) -> DSRActivity:
        activity = await self.repo.get_by_public_id(public_id)
        if not activity:
            raise HTTPException(status_code=404, detail="Activity not found")
        return activity
