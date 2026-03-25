import csv
import io
from typing import List, Optional, Tuple, Dict
from uuid import UUID
from fastapi import HTTPException, status, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
from io import BytesIO

from app.models.dsr_activity_type import DSRActivityType
from app.models.user import User, UserRole
from app.repositories.dsr_activity_type_repository import DSRActivityTypeRepository
from app.schemas.dsr_activity_type import DSRActivityTypeCreate, DSRActivityTypeUpdate


class DSRActivityTypeService:

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = DSRActivityTypeRepository(db)

    async def create_type(
        self,
        data: DSRActivityTypeCreate,
        current_user: User,
    ) -> DSRActivityType:
        # Only admin
        if current_user.role != UserRole.ADMIN:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admins can manage activity types.")

        # Code uniqueness check
        existing = await self.repo.get_by_code(data.code)
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Activity type with code '{data.code}' already exists.",
            )

        obj = DSRActivityType(
            name=data.name,
            code=data.code,
            description=data.description,
            category=data.category,
            is_active=data.is_active,
            sort_order=data.sort_order,
        )
        self.db.add(obj)
        await self.db.flush()
        return obj

    async def update_type(
        self,
        public_id: UUID,
        data: DSRActivityTypeUpdate,
        current_user: User,
    ) -> DSRActivityType:
        if current_user.role != UserRole.ADMIN:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admins can manage activity types.")

        obj = await self.repo.get_by_public_id(public_id)
        if not obj:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Activity type not found.")

        # Code uniqueness check (skip if same object)
        if data.code and data.code != obj.code:
            existing = await self.repo.get_by_code(data.code)
            if existing and existing.id != obj.id:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Activity type with code '{data.code}' already exists.",
                )

        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(obj, field, value)

        await self.db.flush()
        return obj

    async def delete_type(self, public_id: UUID, current_user: User) -> None:
        if current_user.role != UserRole.ADMIN:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admins can manage activity types.")

        obj = await self.repo.get_by_public_id(public_id)
        if not obj:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Activity type not found.")

        obj.soft_delete()
        await self.db.flush()

    async def get_type(self, public_id: UUID) -> DSRActivityType:
        obj = await self.repo.get_by_public_id(public_id)
        if not obj:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Activity type not found.")
        return obj

    async def get_types(
        self,
        skip: int = 0,
        limit: int = 100,
        active_only: bool = False,
        search: Optional[str] = None,
    ) -> Tuple[List[DSRActivityType], int]:
        return await self.repo.get_multi_paginated(
            skip=skip,
            limit=limit,
            active_only=active_only,
            search=search,
        )

    async def get_active_types(self) -> List[DSRActivityType]:
        """Shortcut used by the DSR form dropdown."""
        return await self.repo.get_active_types()

    async def import_activity_types(self, file: UploadFile, current_user: User) -> dict:
        """Import DSR activity types from CSV/Excel: name,code,category,description,sort_order"""
        if current_user.role != UserRole.ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only Admins can import activity types.",
            )

        content = await file.read()
        filename = file.filename.lower()
        rows = []

        if filename.endswith('.csv'):
            try:
                decoded = content.decode("utf-8")
            except UnicodeDecodeError:
                try:
                    decoded = content.decode("utf-8-sig")
                except UnicodeDecodeError:
                    decoded = content.decode("cp1252", errors="replace")
            
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
                sheet = wb.active
                
                # Get headers
                headers = [cell.value for cell in sheet[1]]
                
                # Convert rows to dicts
                for row_idx in range(2, sheet.max_row + 1):
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        if header:
                            val = sheet.cell(row=row_idx, column=col_idx + 1).value
                            row_data[str(header).strip().lower()] = val
                    if any(row_data.values()): # Skip empty rows
                        rows.append(row_data)
            except Exception as e:
                 raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to parse Excel file: {str(e)}")
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported file format. Please use .csv or .xlsx")

        total_rows = len(rows)
        created_count = 0
        skipped_count = 0
        errors = []

        for index, row in enumerate(rows):
            line_num = index + 2 # Header is row 1
            try:
                # Support both case-sensitive and case-insensitive headers from Excel
                def get_val(key):
                    return str(row.get(key) or row.get(key.lower()) or "").strip()

                name = get_val("name")
                code = get_val("code")
                
                if not name or not code:
                    skipped_count += 1
                    errors.append({"row": line_num, "error": "Missing name or code"})
                    continue

                category = get_val("category") or None
                description = get_val("description") or None
                sort_order_str = get_val("sort_order") or "0"
                try:
                    sort_order = int(float(sort_order_str)) # Handle float strings from Excel
                except ValueError:
                    sort_order = 0

                existing = await self.repo.get_by_code(code)
                if existing:
                    # Update existing
                    existing.name = name
                    existing.category = category
                    existing.description = description
                    existing.sort_order = sort_order
                    skipped_count += 1
                else:
                    # Create new
                    await self.repo.create({
                        "name": name,
                        "code": code,
                        "category": category,
                        "description": description,
                        "sort_order": sort_order,
                        "is_active": True
                    })
                    created_count += 1

                await self.db.flush()
            except Exception as e:
                skipped_count += 1
                errors.append({"row": line_num, "error": str(e)})

        return {
            "total_rows": total_rows,
            "created": created_count,
            "skipped": skipped_count,
            "errors": errors
        }

    async def bulk_delete_types(self, public_ids: List[UUID], current_user: User) -> int:
        """Bulk delete activity types (Admin only)."""
        if current_user.role != UserRole.ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only Admins can bulk delete activity types.",
            )
        
        count = await self.repo.bulk_delete(public_ids)
        await self.db.flush()
        return count
