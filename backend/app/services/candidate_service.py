"""Candidate Service"""

import io
from datetime import datetime, date
from typing import List, Optional, Any
from uuid import UUID
from fastapi import HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.candidate import Candidate
from app.models.user import User, UserRole
from app.models.candidate_assignment import CandidateAssignment
from app.schemas.candidate import CandidateCreate, CandidateUpdate
from app.schemas.candidate_assignment import CandidateAssignmentCreate
from app.repositories.candidate_repository import CandidateRepository
from app.services.pincode_service import get_pincode_details
from app.utils.email import send_email, send_export_email


class CandidateService:
    """Service for candidate business logic"""
    
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repository = CandidateRepository(db)

    async def validate_personal_info(self, email: str, phone: str, pincode: str, country_code: str = "IN", exclude_public_id: Optional[UUID] = None) -> dict:
        """Validate email, phone availability and pincode existence"""
        # Check existing email
        existing_email = await self.repository.get_by_email(email)
        if email and existing_email:
            if not exclude_public_id or existing_email.public_id != exclude_public_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Email already registered"
                )
        
        # Check existing phone
        existing_phone = await self.repository.get_by_phone(phone)
        if phone and existing_phone:
            if not exclude_public_id or existing_phone.public_id != exclude_public_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Phone number already registered"
                )

        # Fetch address details from pincode - gracefully handle invalid ones
        try:
            return await get_pincode_details(pincode, country_code=country_code)
        except HTTPException:
            # If invalid pincode, return empty details to allow manual entry
            return {}

    async def create_candidate(self, candidate_in: CandidateCreate) -> Candidate:
        """Create a new candidate with automated address fetch"""
        
        # Validate personal info and get address details
        # Extract country_code if available in candidate_in (will be added to schema)
        country_code = getattr(candidate_in, "country_code", "IN")
        address_details = await self.validate_personal_info(
            candidate_in.email, 
            candidate_in.phone, 
            candidate_in.pincode,
            country_code=country_code
        )
        
        # Prepare data
        candidate_data = candidate_in.model_dump()
        
        # Pop country_code if it is not in the database model to avoid TypeError at repository.create
        candidate_data.pop("country_code", None)
        
        # Handle nested objects -> convert to dict/list for JSON storage
        # (Pydantic model_dump already does this recursively by default if we use mode='json', 
        # but sqlalchemy expects dicts for JSON columns. candidate_in.education_details is a model)
        
        # Override address fields only if they are not manually provided
        candidate_data["city"] = candidate_in.city or address_details.get("city")
        candidate_data["district"] = candidate_in.district or address_details.get("district")
        candidate_data["state"] = candidate_in.state or address_details.get("state")
        
        # JSON fields handling
        if candidate_in.education_details:
             candidate_data["education_details"] = candidate_in.education_details.model_dump()
             
        if candidate_in.disability_details:
             candidate_data["disability_details"] = candidate_in.disability_details.model_dump()

        # Set default registration type in 'other' JSON field
        if not candidate_data.get("other"):
            candidate_data["other"] = {"registration_type": "Registered"}
        elif "registration_type" not in candidate_data["other"]:
            candidate_data["other"]["registration_type"] = "Registered"

        # guardian_details and work_experience are passed as dicts currently

        # Build candidate object (UUID is automatically generated)
        candidate = await self.repository.create(candidate_data)
        
        # Refresh to get the candidate with relationships loaded
        # This ensures the response includes screening, documents, counseling (even if empty)
        candidate_with_details = await self.repository.get_by_public_id_with_details(candidate.public_id)
        
        return candidate_with_details

    async def get_candidate(self, public_id: UUID, with_details: bool = False) -> Candidate:
        """Get candidate by public_id (UUID)"""
        if with_details:
            candidate = await self.repository.get_by_public_id_with_details(public_id)
        else:
            candidate = await self.repository.get_by_public_id(public_id)
            
        if not candidate:
            raise HTTPException(status_code=404, detail="Candidate not found")
        return candidate

    async def get_candidates(
        self, 
        skip: int = 0, 
        limit: int = 100, 
        search: Optional[str] = None, 
        sort_by: Optional[str] = None, 
        sort_order: str = "desc",
        disability_types: Optional[list] = None,
        education_levels: Optional[list] = None,
        cities: Optional[list] = None,
        counseling_status: Optional[str] = None,
        is_experienced: Optional[bool] = None,
        screening_status: Optional[str] = None,
        disability_percentages: Optional[list] = None,
        screening_reasons: Optional[list] = None,
        gender: Optional[str] = None,
        year_of_passing: Optional[list] = None,
        year_of_experience: Optional[str] = None,
        currently_employed: Optional[bool] = None,
        extra_filters: Optional[dict] = None,
        registration_type: Optional[str] = None,
        current_user: Optional[User] = None,
        is_global: bool = False
    ) -> dict:
        """Get list of candidates with total count, supporting optional search, filters, and sorting"""
        
        # Determine assigned_to_id based on user role
        assigned_to_id = None
        if current_user and current_user.role == UserRole.SOURCING and not is_global:
            assigned_to_id = current_user.id
            
        items, total = await self.repository.get_multi(
            skip=skip, 
            limit=limit, 
            search=search, 
            sort_by=sort_by, 
            sort_order=sort_order,
            disability_types=disability_types,
            education_levels=education_levels,
            cities=cities,
            counseling_status=counseling_status,
            is_experienced=is_experienced,
            screening_status=screening_status,
            disability_percentages=disability_percentages,
            screening_reasons=screening_reasons,
            gender=gender,
            year_of_passing=year_of_passing,
            year_of_experience=year_of_experience,
            currently_employed=currently_employed,
            assigned_to_id=assigned_to_id,
            extra_filters=extra_filters,
            registration_type=registration_type
        )
        return {"items": items, "total": total}


    async def update_candidate(self, public_id: UUID, candidate_in: CandidateUpdate) -> Candidate:
        """Update candidate by public_id"""
        candidate = await self.get_candidate(public_id)
        
        update_data = candidate_in.model_dump(exclude_unset=True)
        country_code = update_data.pop("country_code", "IN")
        
        # If pincode changed, update address
        if "pincode" in update_data and update_data["pincode"] != candidate.pincode:
                # We don't have country_code easily here, but we can try to infer 
                # or just use the existing city/state if provided in update_data
                if not all([update_data.get("city"), update_data.get("district"), update_data.get("state")]):
                    # Only fetch if some fields are missing
                    address_details = await get_pincode_details(update_data["pincode"], country_code=country_code)
                    update_data["city"] = update_data.get("city") or address_details.get("city")
                    update_data["district"] = update_data.get("district") or address_details.get("district")
                    update_data["state"] = update_data.get("state") or address_details.get("state")
            
        # JSON fields handling for updates
        if "education_details" in update_data and update_data["education_details"]:
             if hasattr(update_data["education_details"], "model_dump"):
                update_data["education_details"] = update_data["education_details"].model_dump()
        
        if "disability_details" in update_data and update_data["disability_details"]:
             if hasattr(update_data["disability_details"], "model_dump"):
                update_data["disability_details"] = update_data["disability_details"].model_dump()

        # Use internal id for repository update
        return await self.repository.update(candidate.id, update_data)

    async def delete_candidate(self, public_id: UUID) -> bool:
        """Delete candidate by public_id"""
        candidate = await self.get_candidate(public_id)
        # Use internal id for repository delete
        return await self.repository.delete(candidate.id, soft=False)

    async def get_stats(self) -> dict:
        """Get candidate statistics"""
        return await self.repository.get_stats()

    async def get_screening_stats(self, current_user: Optional[User] = None, is_global: bool = False) -> dict:
        """Get screening statistics with optional assignment filter"""
        assigned_to_id = None
        if current_user and current_user.role == UserRole.SOURCING and not is_global:
            assigned_to_id = current_user.id
            
        return await self.repository.get_screening_stats(assigned_to_id=assigned_to_id)

    async def get_unscreened_candidates(
        self, 
        skip: int = 0, 
        limit: int = 100, 
        search: Optional[str] = None, 
        sort_by: Optional[str] = None, 
        sort_order: str = "desc",
        disability_types: Optional[list] = None,
        education_levels: Optional[list] = None,
        cities: Optional[list] = None,
        screening_status: Optional[str] = None,
        is_experienced: Optional[bool] = None,
        counseling_status: Optional[str] = None,
        current_user: Optional[User] = None,
        is_global: bool = False
    ) -> dict:
        """Get list of candidates without screening records with total count, supporting optional search, filters and sorting"""
        
        assigned_to_id = None
        if current_user and current_user.role == UserRole.SOURCING and not is_global:
            assigned_to_id = current_user.id
            
        items, total = await self.repository.get_unscreened(
            skip=skip, 
            limit=limit, 
            search=search, 
            sort_by=sort_by, 
            sort_order=sort_order,
            disability_types=disability_types,
            education_levels=education_levels,
            cities=cities,
            screening_status=screening_status,
            is_experienced=is_experienced,
            counseling_status=counseling_status,
            assigned_to_id=assigned_to_id
        )
        return {"items": items, "total": total}


    async def get_screened_candidates(
        self, 
        skip: int = 0, 
        limit: int = 100, 
        counseling_status: Optional[str] = None, 
        search: Optional[str] = None, 
        document_status: Optional[str] = None, 
        sort_by: Optional[str] = None, 
        sort_order: str = "desc",
        disability_types: Optional[list] = None,
        education_levels: Optional[list] = None,
        cities: Optional[list] = None,
        screening_status: Optional[str] = None,
        is_experienced: Optional[bool] = None,
        current_user: Optional[User] = None,
        is_global: bool = False
    ) -> dict:
        """Get list of candidates with screening records with total count, supporting optional search, filters, document status filter, and sorting"""
        
        assigned_to_id = None
        if current_user and current_user.role == UserRole.SOURCING and not is_global:
            assigned_to_id = current_user.id
            
        items, total = await self.repository.get_screened(
            skip=skip, 
            limit=limit, 
            counseling_status=counseling_status, 
            search=search, 
            document_status=document_status, 
            sort_by=sort_by, 
            sort_order=sort_order,
            disability_types=disability_types,
            education_levels=education_levels,
            cities=cities,
            screening_status=screening_status,
            is_experienced=is_experienced,
            assigned_to_id=assigned_to_id
        )
        return {"items": items, "total": total}

    async def get_filter_options(self) -> dict:
        """Get all unique values for filterable fields"""
        return await self.repository.get_filter_options()

    async def assign_candidate(
        self, 
        public_id: UUID, 
        assignment_in: CandidateAssignmentCreate, 
        current_user: User
    ) -> CandidateAssignment:
        """Assign a candidate to a sourcing user"""
        # 1. Get candidate
        candidate = await self.get_candidate(public_id)
        
        # 2. Verify target user exists and has SOURCING role
        from app.repositories.user_repository import UserRepository
        user_repo = UserRepository(self.db)
        target_user = await user_repo.get(assignment_in.user_id)
        if not target_user:
            raise HTTPException(status_code=404, detail="Target user not found")
        
        if target_user.role not in [UserRole.SOURCING, UserRole.MANAGER]:
            raise HTTPException(
                status_code=400, 
                detail="Candidates can only be assigned to users with SOURCING or MANAGER role"
            )

        # 3. Check for existing assignment
        from sqlalchemy import select
        stmt = select(CandidateAssignment).where(CandidateAssignment.candidate_id == candidate.id)
        result = await self.db.execute(stmt)
        existing_assignment = result.scalars().first()

        if existing_assignment:
            # Update existing
            existing_assignment.user_id = assignment_in.user_id
            existing_assignment.assigned_by_id = current_user.id
            existing_assignment.assigned_at = datetime.now()
            self.db.add(existing_assignment)
            await self.db.commit()
            await self.db.refresh(existing_assignment)
            return existing_assignment
        else:
            # Create new
            new_assignment = CandidateAssignment(
                candidate_id=candidate.id,
                user_id=assignment_in.user_id,
                assigned_by_id=current_user.id,
                assigned_at=datetime.now()
            )
            self.db.add(new_assignment)
            await self.db.commit()
            await self.db.refresh(new_assignment)
            return new_assignment


    async def export_candidates(
        self,
        current_user: User,
        search: Optional[str] = None,
        sort_by: Optional[str] = None,
        sort_order: str = "desc",
        disability_types: Optional[list] = None,
        education_levels: Optional[list] = None,
        cities: Optional[list] = None,
        counseling_status: Optional[str] = None,
        is_experienced: Optional[bool] = None,
        screening_status: Optional[str] = None,
        disability_percentages: Optional[list] = None,
        screening_reasons: Optional[list] = None,
        gender: Optional[str] = None,
        year_of_passing: Optional[list] = None,
        year_of_experience: Optional[str] = None,
        currently_employed: Optional[bool] = None,
        extra_filters: Optional[dict] = None,
        is_global: bool = False,
        columns: Optional[str] = None
    ) -> bool:
        """Fetch all filtered candidates, generate Excel, and email to user"""
        import json
        
        # 1. Fetch all matching candidates (no limit)
        res = await self.get_candidates(
            skip=0,
            limit=10000, # High limit for export
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
            disability_types=disability_types,
            education_levels=education_levels,
            cities=cities,
            counseling_status=counseling_status,
            is_experienced=is_experienced,
            screening_status=screening_status,
            disability_percentages=disability_percentages,
            screening_reasons=screening_reasons,
            gender=gender,
            year_of_passing=year_of_passing,
            year_of_experience=year_of_experience,
            currently_employed=currently_employed,
            extra_filters=extra_filters,
            current_user=current_user,
            is_global=is_global
        )
        candidates = res["items"]
        
        # 2. Parse columns if provided
        column_defs = []
        if columns:
            try:
                column_defs = json.loads(columns)
            except Exception:
                column_defs = []
        
        if not column_defs:
            # Default columns if none provided
            column_defs = [
                {"id": "name", "label": "Name"},
                {"id": "gender", "label": "Gender"},
                {"id": "email", "label": "Email"},
                {"id": "phone", "label": "Phone"},
                {"id": "city", "label": "City"},
                {"id": "education_details.education_level", "label": "Education"},
                {"id": "disability_details.disability_type", "label": "Disability"},
                {"id": "screening.status", "label": "Screening Status"},
                {"id": "counseling.status", "label": "Counseling Status"},
                {"id": "registration_type", "label": "Registration Source"},
                {"id": "created_at", "label": "Registration Date"}
            ]

        # 3. Generate Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates Report"
        
        # Headers
        for col_num, col_def in enumerate(column_defs, 1):
            cell = ws.cell(row=1, column=col_num, value=col_def["label"])
            cell.font = Font(bold=True)
            
        # Data Rows
        for row_num, c in enumerate(candidates, 2):
            for col_num, col_def in enumerate(column_defs, 1):
                col_id = col_def["id"]
                val = ""
                
                try:
                    # Specific mappings for common IDs from constants.ts and CandidateListResponse
                    if col_id == "registration_type":
                        val = (c.other or {}).get("registration_type", "Registered")
                    elif col_id == "disability_type":
                        val = (c.disability_details or {}).get("disability_type", "")
                    elif col_id == "disability_percentage":
                        val = (c.disability_details or {}).get("disability_percentage", "")
                    elif col_id == "education_level":
                        edu = c.education_details or {}
                        degrees = edu.get("degrees", [])
                        val = degrees[0].get("degree_name") if degrees else ""
                    elif col_id == "year_of_passing":
                        edu = c.education_details or {}
                        degrees = edu.get("degrees", [])
                        val = degrees[0].get("year_of_passing") if degrees else ""
                    elif col_id == "screening_status":
                        val = c.screening.status if c.screening else "Pending"
                    elif col_id == "screening_date":
                        val = c.screening.created_at if c.screening else ""
                    elif col_id == "screened_by_name":
                        val = (c.screening.screened_by.full_name or c.screening.screened_by.username) if c.screening and c.screening.screened_by else ""
                    elif col_id == "counseling_status":
                        val = c.counseling.status if c.counseling else ""
                    elif col_id == "counseling_date":
                        val = c.counseling.counseling_date if c.counseling else ""
                    elif col_id == "counselor_name":
                        val = (c.counseling.counselor.full_name or c.counseling.counselor.username) if c.counseling and c.counseling.counselor else ""
                    elif col_id == "is_experienced":
                        val = (c.work_experience or {}).get("is_experienced", False)
                    elif col_id == "year_of_experience":
                        val = (c.work_experience or {}).get("year_of_experience", "")
                    elif col_id == "currently_employed":
                        val = (c.work_experience or {}).get("currently_employed", False)
                    elif col_id == "suitable_job_roles":
                        val = (c.counseling.others or {}).get("suitable_job_roles", []) if c.counseling else []
                    elif col_id == "source_of_info":
                        val = (c.screening.others or {}).get("source_of_info", "") if c.screening else ""
                    elif col_id == "family_annual_income":
                        val = (c.screening.others or {}).get("family_annual_income", "") if c.screening else ""
                    elif col_id == "screening_comments":
                        val = (c.screening.others or {}).get("reason", "") if c.screening else ""
                    elif col_id == "skills" and c.counseling:
                        val = c.counseling.skills
                    elif col_id == "workexperience" and c.counseling:
                        val = c.counseling.workexperience
                    elif col_id == "questions" and c.counseling:
                        val = c.counseling.questions
                    elif col_id.startswith("screening_others."):
                        field_name = col_id.replace("screening_others.", "")
                        val = (c.screening.others or {}).get(field_name, "") if c.screening else ""
                    elif col_id.startswith("counseling_others."):
                        field_name = col_id.replace("counseling_others.", "")
                        val = (c.counseling.others or {}).get(field_name, "") if c.counseling else ""
                    elif "." in col_id:
                        parts = col_id.split(".")
                        obj = c
                        for part in parts:
                            if obj is None: break
                            if isinstance(obj, dict): obj = obj.get(part, "")
                            else: obj = getattr(obj, part, None)
                        val = obj
                    else:
                        val = getattr(c, col_id, "")
                        if (val is None or val == "") and c.screening and hasattr(c.screening, col_id):
                            val = getattr(c.screening, col_id, "")
                        if (val is None or val == "") and c.counseling and hasattr(c.counseling, col_id):
                            val = getattr(c.counseling, col_id, "")
                    
                    # Formatting
                    if isinstance(val, (datetime, date)):
                        val = val.strftime('%Y-%m-%d')
                    elif isinstance(val, bool):
                        val = "Yes" if val else "No"
                    elif val is None:
                        val = ""
                    elif isinstance(val, (dict, list)):
                        if col_id == "family_details" and isinstance(val, list):
                            val = "; ".join([f"{f.get('relation')}: {f.get('name')} ({f.get('occupation', 'N/A')})" for f in val])
                        elif col_id == "skills" and isinstance(val, list):
                            val = ", ".join([f"{s.get('name')} ({s.get('level')})" for s in val])
                        elif col_id == "workexperience" and isinstance(val, list):
                            val = ", ".join([f"{w.get('job_title')} at {w.get('company')}" for w in val])
                        elif col_id == "questions" and isinstance(val, list):
                            val = " | ".join([f"Q: {q.get('question')} A: {q.get('answer')}" for q in val])
                        else:
                            val = json.dumps(val)
                except Exception:
                    val = "Error"
                
                ws.cell(row=row_num, column=col_num, value=str(val) if val is not None else "")

        # Save to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 3. Send Email
        report_name = "Candidates Report"
        filename = f"Candidates_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return await send_export_email(
            to_email=current_user.email,
            user_name=current_user.full_name or current_user.username,
            report_name=report_name,
            file_content=excel_buffer.getvalue(),
            filename=filename
        )
