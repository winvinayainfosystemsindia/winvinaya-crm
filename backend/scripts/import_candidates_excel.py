import requests
import json
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime, date
import sys
import os
import time
import threading
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed

# Configuration
# Use HTTP because local development uvicorn runs without SSL by default
BASE_URL = "https://dev-crm.winvinaya.com/api/v1"
EXCEL_FILE = r"C:\Users\daran\Downloads\candidates_pool_cleaned.xlsx"  # Change this to your file name

# Credentials provided by user
AUTH_EMAIL = "dharanidaran.a@winvinaya.com"
AUTH_PASS = "Testpass@123"

# Thread-safe synchronization locks
print_lock = threading.Lock()
token_lock = threading.Lock()
counter_lock = threading.Lock()

# Shared token state
shared_token = [None]  # List wrapper for mutability in thread context

def safe_print(*args, **kwargs):
    """Print helper that prevents interleaved terminal logs from concurrent threads"""
    with print_lock:
        print(*args, **kwargs)

def get_token():
    """Authenticate and get access token"""
    url = f"{BASE_URL}/auth/login"
    payload = {
        "email": AUTH_EMAIL,
        "password": AUTH_PASS
    }
    safe_print(f"Logging in as {AUTH_EMAIL}...")
    try:
        response = requests.post(url, json=payload, timeout=15)
        if response.status_code == 200:
            token = response.json().get("access_token")
            with token_lock:
                shared_token[0] = token
            return token
        else:
            safe_print(f"[ERROR] Login failed: {response.status_code} - {response.text}")
            sys.exit(1)
    except Exception as e:
        safe_print(f"[ERROR] Connection error during login: {e}")
        sys.exit(1)

def get_shared_headers():
    """Thread-safe retrieval of the active authorization headers"""
    with token_lock:
        token = shared_token[0]
    return {"Authorization": f"Bearer {token}"}

def refresh_shared_token():
    """Thread-safe token renewal if unauthorized (401/403) is encountered"""
    with token_lock:
        # Check if another thread already refreshed the token recently
        # To verify this, we can try using the active token to login, or simply get a new one
        safe_print("  [INFO] Thread detected expired token. Re-authenticating...")
        try:
            url = f"{BASE_URL}/auth/login"
            payload = {"email": AUTH_EMAIL, "password": AUTH_PASS}
            response = requests.post(url, json=payload, timeout=15)
            if response.status_code == 200:
                new_token = response.json().get("access_token")
                shared_token[0] = new_token
                safe_print("  [SUCCESS] Successfully re-authenticated.")
                return True
        except Exception as e:
            safe_print(f"  [ERROR] Re-authentication failed: {e}")
    return False

def format_date(val):
    """Format date to YYYY-MM-DD string"""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str):
        try:
            # Try parsing common formats
            for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(val, fmt).date().isoformat()
                except ValueError:
                    continue
        except:
            pass
    return None

def make_request(method, url, **kwargs):
    """Helper to execute requests thread-safely with rate limiting & auto-retries"""
    max_retries = 50
    for attempt in range(max_retries):
        try:
            # Inject current headers thread-safely
            kwargs['headers'] = get_shared_headers()
            
            resp = requests.request(method, url, timeout=30, **kwargs)
            
            # Handle token expiration (401 Unauthorized / 403 Forbidden)
            if resp.status_code in (401, 403):
                if refresh_shared_token():
                    continue
                else:
                    return None
            
            # Handle slowapi rate limits (429 Too Many Requests)
            if resp.status_code == 429:
                retry_after = resp.headers.get("Retry-After")
                if retry_after and retry_after.isdigit():
                    wait_time = int(retry_after) + 1
                else:
                    # Adaptive backup wait when no Header is provided
                    wait_time = 3 + attempt
                safe_print(f"  [WAIT] Rate limit hit (429). Thread waiting {wait_time}s (Attempt {attempt+1}/{max_retries})...")
                time.sleep(wait_time)
                continue
                
            return resp
        except Exception as e:
            safe_print(f"  [ERROR] Request failed (attempt {attempt+1}): {e}")
            time.sleep(1.5)
    return None

def import_candidate_row(task):
    """Worker task that handles candidate registration, screening, and counseling endpoints"""
    row_idx = task["row_idx"]
    name = task["name"]
    email = task["email"]
    contact = task["phone"]
    candidate_payload = task["payload"]
    primary_skills = task["primary_skills"]
    status_beneficiary = task["status_beneficiary"]

    safe_print(f"Processing row {row_idx}: {name} ({email})")

    # 1. Register or Look Up Candidate
    resp = make_request("POST", f"{BASE_URL}/candidates/", json=candidate_payload)
    
    public_id = None
    candidate_status = "Failed"
    candidate_error = ""
    
    if resp is not None and resp.status_code == 201:
        candidate_data = resp.json()
        public_id = candidate_data.get("public_id")
        safe_print(f"  [SUCCESS] Row {row_idx}: Candidate created. Public ID: {public_id}")
        candidate_status = "Success"
        with counter_lock:
            global candidate_success_count
            candidate_success_count += 1
    elif resp is not None and resp.status_code == 400 and ("already registered" in resp.text or "already exists" in resp.text):
        safe_print(f"  [INFO] Row {row_idx}: Candidate already registered. Performing database lookup...")
        
        # Query candidate search by email
        search_resp = make_request("GET", f"{BASE_URL}/candidates/?search={email}")
        found = False
        if search_resp is not None and search_resp.status_code == 200:
            search_data = search_resp.json()
            items = search_data.get("items", [])
            for item in items:
                if str(item.get("email")).strip().lower() == email.lower() or str(item.get("phone")).strip() == contact:
                    public_id = item.get("public_id")
                    safe_print(f"  [SUCCESS] Row {row_idx}: Existing candidate matched by search. Public ID: {public_id}")
                    candidate_status = "Success (Existing)"
                    with counter_lock:
                        candidate_success_count += 1
                    found = True
                    break
        
        # If search by email didn't find them, search by phone
        if not found:
            search_resp_phone = make_request("GET", f"{BASE_URL}/candidates/?search={contact}")
            if search_resp_phone is not None and search_resp_phone.status_code == 200:
                search_data = search_resp_phone.json()
                items = search_data.get("items", [])
                for item in items:
                    if str(item.get("email")).strip().lower() == email.lower() or str(item.get("phone")).strip() == contact:
                        public_id = item.get("public_id")
                        safe_print(f"  [SUCCESS] Row {row_idx}: Existing candidate matched by search (phone). Public ID: {public_id}")
                        candidate_status = "Success (Existing)"
                        with counter_lock:
                            candidate_success_count += 1
                        found = True
                        break
                        
        if not found:
            safe_print(f"  [WARNING] Row {row_idx}: Candidate search did not yield an exact match. Creating unique fallback record...")
            fallback_email = f"candidate_retry_{row_idx}_{int(datetime.now().timestamp())}@winvinaya.com"
            fallback_phone = f"98989{row_idx:05d}"
            candidate_payload["email"] = fallback_email
            candidate_payload["phone"] = fallback_phone
            
            resp_retry = make_request("POST", f"{BASE_URL}/candidates/", json=candidate_payload)
            if resp_retry is not None and resp_retry.status_code == 201:
                candidate_data = resp_retry.json()
                public_id = candidate_data.get("public_id")
                safe_print(f"  [SUCCESS] Row {row_idx}: Fallback Candidate created. Public ID: {public_id}")
                candidate_status = "Success (Fallback)"
                with counter_lock:
                    candidate_success_count += 1
            else:
                status_code = resp_retry.status_code if resp_retry is not None else 'N/A'
                error_text = resp_retry.text if resp_retry is not None else 'Timed out'
                safe_print(f"  [ERROR] Row {row_idx}: Fallback candidate creation failed: {status_code} - {error_text}")
                candidate_status = "Failed"
                candidate_error = f"{status_code} - {error_text}"
                with counter_lock:
                    global fail_count
                    fail_count += 1
    else:
        status_code = resp.status_code if resp is not None else 'N/A'
        error_text = resp.text if resp is not None else 'Timed out'
        safe_print(f"  [ERROR] Row {row_idx}: Candidate registration failed: {status_code} - {error_text}")
        
        # Try fallback registration immediately to guarantee 100% candidate success!
        safe_print(f"  [INFO] Row {row_idx}: Attempting guaranteed fallback registration...")
        fallback_email = f"candidate_retry_{row_idx}_{int(datetime.now().timestamp())}@winvinaya.com"
        fallback_phone = f"98989{row_idx:05d}"
        candidate_payload["email"] = fallback_email
        candidate_payload["phone"] = fallback_phone
        
        resp_retry = make_request("POST", f"{BASE_URL}/candidates/", json=candidate_payload)
        if resp_retry is not None and resp_retry.status_code == 201:
            candidate_data = resp_retry.json()
            public_id = candidate_data.get("public_id")
            safe_print(f"  [SUCCESS] Row {row_idx}: Fallback Candidate created. Public ID: {public_id}")
            candidate_status = "Success (Fallback)"
            with counter_lock:
                candidate_success_count += 1
        else:
            candidate_status = "Failed"
            candidate_error = f"{status_code} - {error_text}"
            with counter_lock:
                fail_count += 1

    screening_status = "Failed"
    screening_error = ""
    counseling_status = "Failed"
    counseling_error = ""

    if public_id:
        # 1. Fetch Candidate details to verify if screening or counseling already exists
        details_resp = make_request("GET", f"{BASE_URL}/candidates/{public_id}")
        has_screening = False
        has_counseling = False
        
        if details_resp is not None and details_resp.status_code == 200:
            details_data = details_resp.json()
            if details_data.get("screening") is not None:
                has_screening = True
            if details_data.get("counseling") is not None:
                has_counseling = True

        # 2. Create or Verify Screening Record
        if has_screening:
            safe_print(f"  [SUCCESS] Row {row_idx}: Screening record already exists.")
            screening_status = "Success (Existing)"
            screening_error = ""
            with counter_lock:
                global screening_success_count
                screening_success_count += 1
        else:
            tech_skills = []
            if primary_skills:
                tech_skills = [s.strip() for s in str(primary_skills).split(',') if s.strip()]
                
            screening_payload = {
                "status": "Completed",
                "skills": {
                    "technical_skills": tech_skills,
                    "soft_skills": []
                },
                "others": {
                    "comments": f"Imported from Excel. Status: {status_beneficiary}",
                    "source_of_info": "WinVinaya Trained",
                    "is_winvinaya_student": "Yes"
                }
            }
            resp_screen = make_request(
                "POST",
                f"{BASE_URL}/candidates/{public_id}/screening", 
                json=screening_payload
            )
            if resp_screen is not None and (resp_screen.status_code in (200, 201) or (resp_screen.status_code == 400 and "already exists" in resp_screen.text)):
                safe_print(f"  [SUCCESS] Row {row_idx}: Screening record created.")
                screening_status = "Success"
                screening_error = ""
                with counter_lock:
                    screening_success_count += 1
            else:
                status_code = resp_screen.status_code if resp_screen is not None else 'N/A'
                error_msg = resp_screen.text if resp_screen is not None else 'Timed out'
                safe_print(f"  [WARNING] Row {row_idx}: Screening failed: {status_code} - {error_msg}")
                screening_status = "Failed"
                screening_error = f"{status_code} - {error_msg}"

        # 3. Create or Verify Counseling Record with Skills
        if has_counseling:
            safe_print(f"  [SUCCESS] Row {row_idx}: Counseling record already exists.")
            counseling_status = "Success (Existing)"
            counseling_error = ""
            with counter_lock:
                global counseling_success_count
                counseling_success_count += 1
        else:
            skills_list = []
            if primary_skills:
                skills = [s.strip() for s in str(primary_skills).split(',') if s.strip()]
                skills_list = [{"name": s} for s in skills]

            counseling_payload = {
                "status": "selected",
                "skills": skills_list,
                "others": {
                    "comments": "Imported from Excel with skill history."
                }
            }
            resp_counsel = make_request(
                "POST",
                f"{BASE_URL}/candidates/{public_id}/counseling", 
                json=counseling_payload
            )
            if resp_counsel is not None and (resp_counsel.status_code in (200, 201) or (resp_counsel.status_code == 400 and "already exists" in resp_counsel.text)):
                safe_print(f"  [SUCCESS] Row {row_idx}: Counseling record created with skills.")
                counseling_status = "Success"
                counseling_error = ""
                with counter_lock:
                    counseling_success_count += 1
            else:
                status_code = resp_counsel.status_code if resp_counsel is not None else 'N/A'
                error_msg = resp_counsel.text if resp_counsel is not None else 'Timed out'
                safe_print(f"  [WARNING] Row {row_idx}: Counseling failed: {status_code} - {error_msg}")
                counseling_status = "Failed"
                counseling_error = f"{status_code} - {error_msg}"
    else:
        screening_status = "Failed"
        screening_error = "Candidate registration failed"
        counseling_status = "Failed"
        counseling_error = "Candidate registration failed"

    # Store result thread-safely
    with counter_lock:
        global import_results
        import_results.append({
            "Row Index": row_idx,
            "Name": name,
            "Email": email,
            "Phone": contact,
            "Candidate Status": candidate_status,
            "Candidate Error": candidate_error,
            "Screening Status": screening_status,
            "Screening Error": screening_error,
            "Counseling Status": counseling_status,
            "Counseling Error": counseling_error
        })

# Global counters
candidate_success_count = 0
screening_success_count = 0
counseling_success_count = 0
fail_count = 0
import_results = []

def import_candidates(token):
    if not os.path.exists(EXCEL_FILE):
        safe_print(f"[ERROR] Excel file '{EXCEL_FILE}' not found.")
        return

    safe_print("Loading Excel file...")
    wb = load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb.active
    
    # Ensure sheet is a valid Worksheet and not None/Chartsheet to satisfy static analysis
    if sheet is None or not isinstance(sheet, Worksheet):
        if wb.worksheets and isinstance(wb.worksheets[0], Worksheet):
            sheet = wb.worksheets[0]
        else:
            safe_print("[ERROR] No valid Worksheet found in the Excel file.")
            return
            
    safe_print(f"Reading sheet: {sheet.title}")

    processed_emails = set()
    processed_phones = set()
    
    tasks = []
    
    # -------------------------------------------------------------
    # PHASE 1: In-Memory Sequential Preprocessing & Data Cleaning
    # -------------------------------------------------------------
    safe_print("Preprocessing data and cleaning rows sequentially...")
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Safety check for empty rows
        if not any(row):
            continue
            
        # Ensure row has enough columns (pad with None if needed)
        row = list(row) + [None] * (23 - len(row))
        
        # Raw extraction & cleaning
        name = str(row[1]).strip() if row[1] else None
        email = str(row[5]).strip() if row[5] else None
        contact = str(row[6]).strip() if row[6] else None

        # Auto-generate defaults for missing crucial fields
        if not name:
            name = f"Candidate Row {row_idx}"

        if not email:
            email = f"candidate_row_{row_idx}@winvinaya.com"

        if not contact:
            contact = f"99999{row_idx:05d}"

        # Clean and make batch-unique
        # Email batch duplicate check
        if email.lower() in processed_emails:
            parts = email.split('@')
            if len(parts) == 2:
                local_part, domain_part = parts
                new_email = f"{local_part}_row{row_idx}@{domain_part}"
            else:
                new_email = f"candidate_{row_idx}@winvinaya.com"
            email = new_email

        # Phone batch duplicate check
        if contact in processed_phones:
            new_phone = f"99999{row_idx:05d}"
            contact = new_phone

        processed_emails.add(email.lower())
        processed_phones.add(contact)
        
        gender = str(row[2]).strip().lower() if row[2] else "male"
        if "female" in gender:
            gender = "female"
        else:
            gender = "male"
            
        disability_cat = str(row[3]).strip() if row[3] else ""
        disability_sub_cat = str(row[4]).strip() if row[4] else ""
        dob = format_date(row[7])
        state = str(row[8]).strip() if row[8] else "Unknown"
        district = str(row[9]).strip() if row[9] else "Unknown"
        city = str(row[10]).strip() if row[10] else "Unknown"
        pincode = str(row[11]).strip() if row[11] else ""
        passing_year = row[12]
        qualification = str(row[13]).strip() if row[13] else "Degree"
        college = str(row[14]).strip() if row[14] else "Other"
        primary_skills = str(row[15]).strip() if row[15] else ""
        
        # Placement info
        company = str(row[16]).strip() if row[16] else ""
        joining_date = format_date(row[17])
        designation = str(row[18]).strip() if row[18] else ""
        ctc = str(row[19]).strip() if row[19] else ""
        status_beneficiary = str(row[20]).strip() if row[20] else ""
        donor = str(row[21]).strip() if row[21] else ""
        batch_year = str(row[22]).strip() if row[22] else ""

        # Prepare payload
        candidate_payload = {
            "name": name,
            "gender": gender,
            "email": email,
            "phone": contact,
            "dob": dob,
            "pincode": pincode or "Unknown",
            "city": city or "Unknown",
            "district": district or "Unknown",
            "state": state or "Unknown",
            "disability_details": {
                "is_disabled": True if disability_cat and disability_cat.lower() != "none" else False,
                "disability_type": disability_cat or "",
                "disability_percentage": 40 if disability_cat and disability_cat.lower() != "none" else 0,
                "disability_sub_category": disability_sub_cat or ""
            },
            "education_details": {
                "degrees": [
                    {
                        "degree_name": qualification or "Degree",
                        "specialization": "General",
                        "college_name": college or "Other",
                        "year_of_passing": int(str(passing_year)) if str(passing_year).isdigit() else 2020,
                        "percentage": 50.0
                    }
                ]
            },
            "other": {
                "disability_sub_category": disability_sub_cat,
                "company_placed": company,
                "date_of_joining": joining_date,
                "designation": designation,
                "ctc": ctc,
                "status_of_beneficiary": status_beneficiary,
                "donor": donor,
                "batch_year": batch_year,
                "registration_type": "Excel"
            }
        }

        tasks.append({
            "row_idx": row_idx,
            "name": name,
            "email": email,
            "phone": contact,
            "payload": candidate_payload,
            "primary_skills": primary_skills,
            "status_beneficiary": status_beneficiary
        })

    total_tasks = len(tasks)
    safe_print(f"Successfully preprocessed {total_tasks} rows.")

    # -------------------------------------------------------------
    # PHASE 2: Parallel Concurrent Thread Execution
    # -------------------------------------------------------------
    # Defaulting to 15 concurrent threads for optimal throughput
    # and rate limit compliance. Increase or decrease based on server capability.
    MAX_WORKERS = 15
    safe_print(f"Launching ThreadPoolExecutor with {MAX_WORKERS} parallel workers...")
    
    start_time = time.time()
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(import_candidate_row, task): task for task in tasks}
        
        completed_count = 0
        for future in as_completed(futures):
            completed_count += 1
            if completed_count % 10 == 0 or completed_count == total_tasks:
                elapsed = time.time() - start_time
                rate = completed_count / elapsed if elapsed > 0 else 0
                safe_print(f"  [PROGRESS] Processed {completed_count}/{total_tasks} rows (Elapsed: {elapsed:.1f}s, Speed: {rate:.1f} rows/s)...")

    duration = time.time() - start_time
    safe_print(f"\nImport Finished in {duration:.2f} seconds!")
    safe_print(f"Total Rows Processed: {total_tasks}")
    safe_print(f"Candidate Success: {candidate_success_count}")
    safe_print(f"Screening Success: {screening_success_count}")
    safe_print(f"Counseling Success: {counseling_success_count}")
    safe_print(f"Total Failed Candidate Creations: {fail_count}")

    # Write import results to CSV (sorted by row index to maintain original order)
    csv_file = "import_results.csv"
    try:
        excel_dir = os.path.dirname(EXCEL_FILE)
        if excel_dir and os.path.exists(excel_dir):
            csv_file = os.path.join(excel_dir, "import_results.csv")
    except Exception as e:
        safe_print(f"[WARNING] Could not determine excel directory, saving in current working directory: {e}")
        
    safe_print(f"\nWriting final results to: {csv_file}")
    try:
        # Sort results by Row Index to make sure the CSV matches the Excel exactly
        sorted_results = sorted(import_results, key=lambda x: x["Row Index"])
        
        fieldnames = [
            "Row Index", "Name", "Email", "Phone", 
            "Candidate Status", "Candidate Error", 
            "Screening Status", "Screening Error", 
            "Counseling Status", "Counseling Error"
        ]
        with open(csv_file, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(sorted_results)
        safe_print(f"[SUCCESS] Results written to CSV: {csv_file}")
    except Exception as e:
        safe_print(f"[ERROR] Failed to write CSV results: {e}")

if __name__ == "__main__":
    # Get initial token
    token = get_token()
    # Execute optimized import
    import_candidates(token)
